import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# Inicialize o driver do Microsoft Edge
driver = webdriver.Edge()

# Abra o site de vendas
url = "https://www.amazon.com.br/gp/browse.html?node=17351089011&ref_=nav_em__pc_gamer_0_2_17_5"
driver.get(url)

# Aguarde um tempo para a página carregar completamente
time.sleep(5)

# extrair todos os títulos
titulos = driver.find_elements(
    By.XPATH, "(//span[@class='a-size-base-plus a-color-base a-text-normal'])"
)
# extrair todos os preços
precos = driver.find_elements(By.XPATH, "(//span[@class='a-price-whole'])")
# Crie um novo arquivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = "Nome do Produto"
sheet['B1'] = "Valor do Produto"

# Escreva os dados na planilha
for i in range(len(titulos)):
    sheet.cell(row=i+2, column=1, value=titulos[i].text)
    sheet.cell(row=i+2, column=2, value=precos[i].text)

# Salve o arquivo Excel
workbook.save("produtos.xlsx")

# Feche o navegador
driver.quit()